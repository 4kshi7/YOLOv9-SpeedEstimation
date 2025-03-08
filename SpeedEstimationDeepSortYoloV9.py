import cv2
import numpy as np
import time
import torch
from models.common import DetectMultiBackend
from utils.general import non_max_suppression, scale_boxes
from utils.torch_utils import select_device
from deep_sort_realtime.deepsort_tracker import DeepSort
from scipy.spatial import distance
import pytesseract
from datetime import datetime
import openpyxl
import threading
from collections import deque
import os
import zipfile


# Configure Tesseract OCR
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

# Select Device
device = select_device(0)  # Force CUDA device selection
print(f"Using device: {device}")

# Load YOLO Model
weights = "weights/yolov9-e.pt"
vehicle_model = DetectMultiBackend(weights, device=device)

# Initialize DeepSORT Tracker with tuned parameters
tracker = DeepSort(max_age=30, n_init=3, max_iou_distance=0.4)  # More strict tracking

# Load Video
video_path = r"C:\Users\acre 2\Downloads\video.MP4"
cap = cv2.VideoCapture(video_path)

if not cap.isOpened():
    print("[ERROR] Video file could not be loaded!")
    exit()

# Get FPS dynamically from the video
fps = cap.get(cv2.CAP_PROP_FPS) or 30
frame_time = 1 / fps  # Time difference per frame

# Speed Settings
speed_limit_kmh = 10

# Data Structures
vehicle_tracks = {}
speed_history = {}

# Excel Log File
excel_file = r"D:\yolov9\yolov9\overspeed_log.xlsx"

# Ensure the Excel file exists before logging data
if not os.path.exists(excel_file):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Timestamp", "Vehicle ID", "Plate Number", "Speed (km/h)"])
    wb.save(excel_file)
    wb.close()
    print("[INFO] Created new Excel file for logging: overspeed_log.xlsx")

# Function to Save Data to Excel
from threading import Lock

excel_lock = Lock()  # Create a Lock to prevent concurrent writes

def save_to_excel(vehicle_id, plate_text, speed_kmh):
    with excel_lock:  # Lock to ensure only one thread writes at a time
        try:
            wb = openpyxl.load_workbook(excel_file)
        except (FileNotFoundError, zipfile.BadZipFile):  # Catch corrupted file error
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Timestamp", "Vehicle ID", "Plate Number", "Speed (km/h)"])
        else:
            ws = wb.active

        # Append Data
        ws.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), vehicle_id, plate_text, round(speed_kmh, 2)])
        
        wb.save(excel_file)
        wb.close()
        
        print(f"[LOG] Overspeeding Vehicle Logged: ID={vehicle_id}, Plate={plate_text}, Speed={round(speed_kmh, 2)} km/h")


# Function to Process License Plate Using Tesseract OCR
def read_plate(plate_roi):
    gray = cv2.cvtColor(plate_roi, cv2.COLOR_BGR2GRAY)
    plate_text = pytesseract.image_to_string(gray, config='--oem 3 --psm 8').strip()
    return plate_text.replace(" ", "").replace("\n", "")

# Multi-threaded License Plate Processing
def process_plate_async(plate_roi, vehicle_id, speed_kmh):
    threading.Thread(target=save_plate, args=(plate_roi, vehicle_id, speed_kmh)).start()

def save_plate(plate_roi, vehicle_id, speed_kmh):
    plate_text = read_plate(plate_roi)
    save_to_excel(vehicle_id, plate_text, speed_kmh)

# Adaptive Frame Skipping
frame_skip_factor = 3
frame_count = 0

while cap.isOpened():
    ret, frame = cap.read()
    if not ret:
        break

    frame_count += 1
    if frame_count % frame_skip_factor != 0:
        continue  # Adaptive skipping to improve efficiency

    img = cv2.resize(frame, (640, 384))
    img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)

    # Convert to Torch Tensor
    img_tensor = torch.from_numpy(img_rgb).to(device)
    img_tensor = img_tensor.float() / 255.0
    img_tensor = img_tensor.permute(2, 0, 1).unsqueeze(0)

    with torch.no_grad():
        pred = vehicle_model(img_tensor)[0]
        pred = non_max_suppression(pred, 0.4, 0.5, classes=None, agnostic=False)

    detections = []
    for det in pred:
        if len(det):
            det = det.cpu().numpy()
            det[:, :4] = scale_boxes(img_tensor.shape[2:], det[:, :4], frame.shape).round()
            for *xyxy, conf, cls in reversed(det):
                x1, y1, x2, y2 = map(int, xyxy)
                detections.append([[x1, y1, x2 - x1, y2 - y1], conf, "vehicle"])

    # Update DeepSORT Tracker
    tracks = tracker.update_tracks(detections, frame=frame)

    for track in tracks:
        if not track.is_confirmed():
            continue

        vehicle_id = track.track_id
        x1, y1, w, h = map(int, track.to_ltrb())
        center_x, center_y = int(x1 + w / 2), int(y1 + h / 2)

        avg_vehicle_length = 4.5  # Meters
        meters_per_pixel = avg_vehicle_length / max(w, 1)

        # Speed Calculation with Kalman Filter Smoothing
        smooth_speed_kmh = 0.0
        if vehicle_id in vehicle_tracks:
            px, py, pt, _ = vehicle_tracks[vehicle_id]
            dist_meters = distance.euclidean((px, py), (center_x, center_y)) * meters_per_pixel
            dt = frame_time  # Use real frame time for better accuracy

            if dt > 0:
                raw_speed_kmh = (dist_meters / dt) * 3.6
                if vehicle_id not in speed_history:
                    speed_history[vehicle_id] = deque(maxlen=5)
                speed_history[vehicle_id].append(raw_speed_kmh)
                smooth_speed_kmh = np.mean(speed_history[vehicle_id])  # Kalman smoothing

                # Log Overspeeding Vehicles
                if smooth_speed_kmh > speed_limit_kmh:
                    plate_roi = frame[y1:y1 + h, x1:x1 + w]
                    if plate_roi.size > 0:
                        process_plate_async(plate_roi, vehicle_id, smooth_speed_kmh)
                    print(f"[ALERT] Overspeeding! Vehicle {vehicle_id} at {round(smooth_speed_kmh, 2)} km/h")

        vehicle_tracks[vehicle_id] = (center_x, center_y, time.time(), smooth_speed_kmh)

        # Draw Tracking Box & Speed
        color = (0, 0, 255) if smooth_speed_kmh > speed_limit_kmh else (0, 255, 0)
        cv2.rectangle(frame, (x1, y1), (x1 + w, y1 + h), color, 2)
        cv2.putText(frame, f"ID {vehicle_id} | {round(smooth_speed_kmh, 2)} km/h", (x1, y1 - 10),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2)

    cv2.imshow("Vehicle Speed Detection", frame)
    if cv2.waitKey(1) & 0xFF == ord("q"):
        break

cap.release()
cv2.destroyAllWindows()
print("[INFO] Detection completed.")
